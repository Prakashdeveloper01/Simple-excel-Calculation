from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
import openpyxl
import os

app = FastAPI()

OUTPUT_FILE = "calculator_output.xlsx"


class InputData(BaseModel):
    value_a: float
    value_b: float


@app.post("/process")
def process_data(data: InputData):
    """
    Appends a new row with Value A, Value B, and Formula C to OUTPUT_FILE.
    """
    try:
        if os.path.exists(OUTPUT_FILE):
            wb = openpyxl.load_workbook(OUTPUT_FILE)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            # Headers
            ws["A1"] = "Value A"
            ws["B1"] = "Value B"
            ws["C1"] = "Total (A+B)"

        # Determine next row (if file empty or just headers, starts at 2)
        next_row = ws.max_row + 1
        if next_row < 2:
            next_row = 2

        # Values
        ws.cell(row=next_row, column=1, value=data.value_a)
        ws.cell(row=next_row, column=2, value=data.value_b)

        # Dynamic Formula
        ws.cell(row=next_row, column=3, value=f"=SUM(A{next_row}, B{next_row})")

        # Save
        try:
            wb.save(OUTPUT_FILE)
        except PermissionError:
            raise HTTPException(
                status_code=409,  # Conflict
                detail=f"Permission denied: '{OUTPUT_FILE}'. Please close the file in Excel and try again.",
            )

        return {
            "message": "Row added successfully.",
            "output_file": OUTPUT_FILE,
            "row_number": next_row,
            "input_values": {"a": data.value_a, "b": data.value_b},
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/read-output")
def read_output():
    """
    Reads ALL rows from OUTPUT_FILE and returns them as a list.
    """
    if not os.path.exists(OUTPUT_FILE):
        raise HTTPException(
            status_code=404, detail="Output file not found. Run /process first."
        )

    try:
        wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
        ws = wb.active

        all_records = []

        # Iterate from row 2 (skipping header) to max_row
        for row in range(2, ws.max_row + 1):
            val_a = ws.cell(row=row, column=1).value
            val_b = ws.cell(row=row, column=2).value

            # Handle cases where read values might be None
            if val_a is None and val_b is None:
                continue  # Skip empty rows

            if val_a is None:
                val_a = 0
            if val_b is None:
                val_b = 0

            # Read Column C (Total) directly from the sheet
            val_c = ws.cell(row=row, column=3).value

            all_records.append(
                {
                    "row": row,
                    "column_a": val_a,
                    "column_b": val_b,
                    "column_c_total": val_c,
                }
            )

        return {
            "file": OUTPUT_FILE,
            "total_records": len(all_records),
            "data": all_records,
            "note": "Values read directly from Excel file.",
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn

    # Only run server if script is executed directly
    uvicorn.run(app, host="127.0.0.1", port=8009)
